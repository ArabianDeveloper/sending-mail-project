import sys
import csv
import json
import smtplib
import base64
import os
import pickle
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QLineEdit, QPushButton, QFileDialog,
    QTableWidget, QTableWidgetItem, QTextEdit, QSpinBox,
    QCheckBox, QComboBox, QProgressBar, QMessageBox, QStatusBar
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QIcon, QColor
import openpyxl

SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def get_credentials():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    return creds

class EmailSenderThread(QThread):
    progress = pyqtSignal(int)
    log = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, recipients, subject, body, sender_email, smtp_server, smtp_port, 
                 email_column, attachment_folder=None, attachment_column=None, use_oauth=False):
        super().__init__()
        self.recipients = recipients
        self.subject = subject
        self.body = body
        self.sender_email = sender_email
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.email_column = email_column
        self.attachment_folder = attachment_folder
        self.attachment_column = attachment_column
        self.use_oauth = use_oauth
        self.total = len(recipients)
        self.sent = 0
        self.failed = 0
        if self.use_oauth:
            self.creds = get_credentials()
            self.service = build('gmail', 'v1', credentials=self.creds)

    def run(self):
        try:
            if self.use_oauth:
                self.log.emit("Using Gmail API with OAuth...")
            else:
                self.log.emit(f"Connecting to {self.smtp_server}:{self.smtp_port}...")
                server = smtplib.SMTP(self.smtp_server, self.smtp_port)
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                self.log.emit("✓ Connected successfully!")

            for idx, recipient in enumerate(self.recipients):
                try:
                    email = str(recipient.get(self.email_column, '')).strip()
                    
                    # Validate email
                    if not email or '@' not in email:
                        self.failed += 1
                        self.log.emit(f"✗ Skipped invalid email: {email}")
                        self.progress.emit(int((idx + 1) / self.total * 100))
                        continue
                    
                    personalized_subject = self.subject
                    personalized_body = self.body

                    # Replace all placeholders with recipient data
                    for key, value in recipient.items():
                        if value is not None:
                            placeholder = f"{{{{{key}}}}}"
                            value_str = str(value).strip()
                            personalized_subject = personalized_subject.replace(placeholder, value_str)
                            personalized_body = personalized_body.replace(placeholder, value_str)

                    # Use multipart/mixed as the root so attachments are preserved
                    msg = MIMEMultipart('mixed')
                    msg['From'] = self.sender_email
                    msg['To'] = email
                    msg['Subject'] = personalized_subject

                    # Create the body part separately to support HTML content
                    body_part = MIMEMultipart('alternative')
                    body_part.attach(MIMEText(personalized_body, 'html'))
                    msg.attach(body_part)
                    
                    # Attach file if attachment column is specified
                    if self.attachment_column and self.attachment_folder:
                        filename = str(recipient.get(self.attachment_column, '')).strip()
                        if filename and not filename.lower().endswith('.pdf'):
                            filename += '.pdf'
                        filename = filename.strip()
                        print(f"Looking for attachment: {filename}")
                        if filename:
                            file_path = Path(self.attachment_folder) / filename
                            if file_path.exists():
                                try:
                                    with open(file_path, 'rb') as attachment:
                                        attachment_part = MIMEBase('application', 'octet-stream')
                                        attachment_part.set_payload(attachment.read())
                                    encoders.encode_base64(attachment_part)
                                    attachment_part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                                    msg.attach(attachment_part)
                                    self.log.emit(f"  📎 Attached: {filename}")
                                except Exception as file_err:
                                    self.log.emit(f"  ⚠️ Could not attach {filename}: {str(file_err)}")
                            else:
                                self.log.emit(f"  ⚠️ File not found: {file_path}")
                                continue
                    
                    if self.use_oauth:
                        raw_msg = base64.urlsafe_b64encode(msg.as_bytes()).decode()
                        message = {'raw': raw_msg}
                        self.service.users().messages().send(userId='me', body=message).execute()
                    else:
                        server.send_message(msg)

                    self.sent += 1
                    self.log.emit(f"✓ Sent to {email}")

                except Exception as e:
                    self.failed += 1
                    email_display = email if email else "[empty]"
                    self.log.emit(f"✗ Failed to send to {email_display}: {str(e)}")

                self.progress.emit(int((idx + 1) / self.total * 100))

            if not self.use_oauth:
                server.quit()
            message = f"Completed! Sent: {self.sent}, Failed: {self.failed}"
            self.finished.emit(True, message)
            self.log.emit(f"\n{message}")

        except Exception as e:
            self.finished.emit(False, str(e))
            self.log.emit(f"✗ Error: {str(e)}")

class EmailSenderApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Bulk Email Sender")
        self.setGeometry(100, 100, 1000, 800)
        self.load_settings()
        self.init_ui()

    def load_settings(self):
        self.settings_file = Path("email_settings.json")
        if self.settings_file.exists():
            with open(self.settings_file, 'r') as f:
                self.settings = json.load(f)
        else:
            self.settings = {}

    def save_settings(self):
        with open(self.settings_file, 'w') as f:
            json.dump(self.settings, f, indent=2)

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()

        # Create tabs
        tabs = QTabWidget()
        
        # Tab 1: File Selection
        file_tab = self.create_file_tab()
        tabs.addTab(file_tab, "📁 Recipients")

        # Tab 2: SMTP Settings
        smtp_tab = self.create_smtp_tab()
        tabs.addTab(smtp_tab, "⚙️ SMTP Settings")

        # Tab 3: Email Composition
        email_tab = self.create_email_tab()
        tabs.addTab(email_tab, "✉️ Email")

        # Tab 4: Send & Logs
        send_tab = self.create_send_tab()
        tabs.addTab(send_tab, "📤 Send")

        main_layout.addWidget(tabs)
        central_widget.setLayout(main_layout)
        self.setStatusBar(QStatusBar())

    def create_file_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # File upload
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("Select File:"))
        self.file_path_input = QLineEdit()
        self.file_path_input.setReadOnly(True)
        file_layout.addWidget(self.file_path_input)
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)

        # Email column selection
        col_layout = QHBoxLayout()
        col_layout.addWidget(QLabel("Email Column:"))
        self.email_column = QComboBox()
        col_layout.addWidget(self.email_column)
        layout.addLayout(col_layout)

        # Attachment section
        layout.addWidget(QLabel("📎 Attachments (Optional):"))
        
        # Attachment folder
        attachment_folder_layout = QHBoxLayout()
        attachment_folder_layout.addWidget(QLabel("PDF Folder:"))
        self.attachment_folder_input = QLineEdit()
        self.attachment_folder_input.setReadOnly(True)
        self.attachment_folder_input.setPlaceholderText("Select folder containing PDF files")
        attachment_folder_layout.addWidget(self.attachment_folder_input)
        browse_folder_btn = QPushButton("Browse Folder")
        browse_folder_btn.clicked.connect(self.browse_attachment_folder)
        attachment_folder_layout.addWidget(browse_folder_btn)
        layout.addLayout(attachment_folder_layout)

        # Attachment column selection
        attachment_col_layout = QHBoxLayout()
        attachment_col_layout.addWidget(QLabel("PDF Filename Column:"))
        self.attachment_column = QComboBox()
        attachment_col_layout.addWidget(self.attachment_column)
        layout.addLayout(attachment_col_layout)

        # Preview table
        layout.addWidget(QLabel("Preview Recipients:"))
        self.preview_table = QTableWidget()
        self.preview_table.setMaximumHeight(300)
        layout.addWidget(self.preview_table)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_smtp_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # Preset selection
        preset_layout = QHBoxLayout()
        preset_layout.addWidget(QLabel("Preset:"))
        self.preset = QComboBox()
        self.preset.addItems(["Custom", "Gmail", "Outlook", "Office 365"])
        self.preset.currentTextChanged.connect(self.on_preset_changed)
        preset_layout.addWidget(self.preset)
        preset_layout.addStretch()
        layout.addLayout(preset_layout)

        # Email
        email_layout = QHBoxLayout()
        email_layout.addWidget(QLabel("Email:"))
        self.smtp_email = QLineEdit()
        self.smtp_email.setText(self.settings.get('smtp_email', ''))
        email_layout.addWidget(self.smtp_email)
        layout.addLayout(email_layout)

        # Password
        pwd_layout = QHBoxLayout()
        pwd_layout.addWidget(QLabel("Password:"))
        self.smtp_password = QLineEdit()
        self.smtp_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.smtp_password.setText(self.settings.get('smtp_password', ''))
        pwd_layout.addWidget(self.smtp_password)
        layout.addLayout(pwd_layout)

        password_help = QLabel(
            'If you are using Gmail SMTP with 2-Step Verification enabled, create an app password here: '
            '<a href="https://myaccount.google.com/apppasswords">Google App Passwords</a>'
        )
        password_help.setOpenExternalLinks(True)
        password_help.setStyleSheet("color: #0066cc; font-style: italic;")
        layout.addWidget(password_help)

        # OAuth Checkbox
        oauth_layout = QHBoxLayout()
        self.use_oauth = QCheckBox("Use OAuth2 for Gmail")
        self.use_oauth.setChecked(self.settings.get('use_oauth', False))
        oauth_layout.addWidget(self.use_oauth)
        oauth_layout.addStretch()
        layout.addLayout(oauth_layout)

        # SMTP Server
        server_layout = QHBoxLayout()
        server_layout.addWidget(QLabel("SMTP Server:"))
        self.smtp_server = QLineEdit()
        self.smtp_server.setText(self.settings.get('smtp_server', 'smtp.gmail.com'))
        server_layout.addWidget(self.smtp_server)
        layout.addLayout(server_layout)

        # SMTP Port
        port_layout = QHBoxLayout()
        port_layout.addWidget(QLabel("SMTP Port:"))
        self.smtp_port = QSpinBox()
        self.smtp_port.setValue(self.settings.get('smtp_port', 587))
        self.smtp_port.setRange(1, 65535)
        port_layout.addWidget(self.smtp_port)
        port_layout.addStretch()
        layout.addLayout(port_layout)

        # Test connection
        test_btn = QPushButton("Test Connection")
        test_btn.clicked.connect(self.test_connection)
        layout.addWidget(test_btn)

        # Save settings
        save_btn = QPushButton("Save Settings")
        save_btn.clicked.connect(self.save_smtp_settings)
        layout.addWidget(save_btn)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_email_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # Subject
        layout.addWidget(QLabel("Subject:"))
        self.email_subject = QLineEdit()
        self.email_subject.setText(self.settings.get('email_subject', 'Hello {{Name}}'))
        layout.addWidget(self.email_subject)

        # Body
        layout.addWidget(QLabel("Email Body (HTML):"))
        self.email_body = QTextEdit()
        self.email_body.setMinimumHeight(300)
        self.email_body.setText(self.settings.get('email_body', '<p>Hi {{Name}},</p><p>This is a test email.</p>'))
        layout.addWidget(self.email_body)

        # Help text
        help_text = QLabel("💡 Use {{ColumnName}} for personalization (e.g., {{Name}}, {{Email}}, {{Department}})")
        help_text.setStyleSheet("color: #666; font-style: italic;")
        layout.addWidget(help_text)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_send_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # Progress bar
        layout.addWidget(QLabel("Progress:"))
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        # Logs
        layout.addWidget(QLabel("Logs:"))
        self.logs_text = QTextEdit()
        self.logs_text.setReadOnly(True)
        layout.addWidget(self.logs_text)

        # Send button
        self.send_btn = QPushButton("Send Emails")
        self.send_btn.setMinimumHeight(50)
        font = self.send_btn.font()
        font.setPointSize(12)
        font.setBold(True)
        self.send_btn.setFont(font)
        self.send_btn.clicked.connect(self.send_emails)
        layout.addWidget(self.send_btn)

        widget.setLayout(layout)
        return widget

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select CSV/Excel File", "", "CSV/Excel Files (*.csv *.xlsx);;All Files (*)"
        )
        if file_path:
            self.file_path_input.setText(file_path)
            self.load_file_columns(file_path)

    def browse_attachment_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Folder with PDF Files")
        if folder_path:
            self.attachment_folder_input.setText(folder_path)

    def load_file_columns(self, file_path):
        try:
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    columns = reader.fieldnames or []
            else:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                columns = [cell.value for cell in ws[1]]

            self.email_column.clear()
            self.email_column.addItems([col for col in columns if col])
            
            # Also populate attachment column dropdown
            self.attachment_column.clear()
            self.attachment_column.addItem("(None)")
            self.attachment_column.addItems([col for col in columns if col])
            
            self.preview_recipients(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load file: {str(e)}")

    def preview_recipients(self, file_path):
        try:
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            else:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

            self.preview_table.setColumnCount(len(rows[0]) if rows else 0)
            self.preview_table.setRowCount(len(rows))
            self.preview_table.setHorizontalHeaderLabels(rows[0].keys() if rows else [])

            for row_idx, row in enumerate(rows):
                for col_idx, (key, value) in enumerate(row.items()):
                    self.preview_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to preview file: {str(e)}")

    def on_preset_changed(self, preset):
        presets = {
            "Gmail": ("smtp.gmail.com", 587),
            "Outlook": ("smtp-mail.outlook.com", 587),
            "Office 365": ("smtp.office365.com", 587),
        }
        if preset in presets:
            self.smtp_server.setText(presets[preset][0])
            self.smtp_port.setValue(presets[preset][1])

    def test_connection(self):
        if self.use_oauth.isChecked():
            QMessageBox.information(self, "OAuth Selected", "OAuth uses Gmail API, no SMTP test needed. Credentials are tested during send.")
            return
        try:
            server = smtplib.SMTP(self.smtp_server.text(), self.smtp_port.value())
            server.starttls()
            server.login(self.smtp_email.text(), self.smtp_password.text())
            server.quit()
            QMessageBox.information(self, "Success", "Connection successful!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Connection failed: {str(e)}")

    def save_smtp_settings(self):
        self.settings['smtp_email'] = self.smtp_email.text()
        self.settings['smtp_password'] = self.smtp_password.text()
        self.settings['smtp_server'] = self.smtp_server.text()
        self.settings['smtp_port'] = self.smtp_port.value()
        self.settings['use_oauth'] = self.use_oauth.isChecked()
        self.settings['email_subject'] = self.email_subject.text()
        self.settings['email_body'] = self.email_body.toPlainText()
        self.save_settings()
        QMessageBox.information(self, "Success", "Settings saved!")

    def send_emails(self):
        
        if not self.file_path_input.text():
            QMessageBox.warning(self, "Error", "Please select a file first")
            return

        email_col = self.email_column.currentText()
        if not email_col:
            QMessageBox.warning(self, "Error", "Please select an email column")
            return

        try:
            recipients = self.load_recipients(self.file_path_input.text())
            if not recipients:
                QMessageBox.warning(self, "Error", "No valid recipients found")
                return

            self.send_btn.setEnabled(False)
            self.logs_text.clear()
            self.progress_bar.setValue(0)

            # Get attachment settings
            attachment_folder = self.attachment_folder_input.text() if self.attachment_folder_input.text() else None
            attachment_column = self.attachment_column.currentText()
            if attachment_column == "(None)":
                attachment_column = None

            self.sender_thread = EmailSenderThread(
                recipients,
                self.email_subject.text(),
                self.email_body.toPlainText(),
                self.smtp_email.text(),
                self.smtp_server.text(),
                self.smtp_port.value(),
                email_col,
                attachment_folder,
                attachment_column,
                self.use_oauth.isChecked()
            )
            self.sender_thread.progress.connect(self.progress_bar.setValue)
            self.sender_thread.log.connect(self.append_log)
            self.sender_thread.finished.connect(self.on_send_finished)
            self.sender_thread.start()

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            self.send_btn.setEnabled(True)

    def load_recipients(self, file_path):
        recipients = []
        try:
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    recipients = list(reader)
            else:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                headers = [cell.value for cell in ws[1] if cell.value]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row[:len(headers)]))
                    recipients.append(row_dict)

            email_col = self.email_column.currentText()
            
            # Filter: only keep recipients with valid, non-empty emails
            valid_recipients = []
            for r in recipients:
                email = r.get(email_col, '').strip() if r.get(email_col) else ''
                if email and '@' in email:  # Basic email validation
                    valid_recipients.append(r)
            
            self.append_log(f"Loaded {len(valid_recipients)} valid recipients from {len(recipients)} total rows")
            return valid_recipients
        except Exception as e:
            self.append_log(f"Error loading recipients: {str(e)}")
            return []

    def append_log(self, message):
        self.logs_text.append(message)

    def on_send_finished(self, success, message):
        self.send_btn.setEnabled(True)
        if success:
            QMessageBox.information(self, "Success", message)
        else:
            QMessageBox.critical(self, "Error", message)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = EmailSenderApp()
    window.show()
    sys.exit(app.exec())